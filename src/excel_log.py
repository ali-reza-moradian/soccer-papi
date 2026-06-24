"""Append every arb found (REAL and SHADOW) to data/arbs_log.xlsx — one row per arb per scan.

A human-friendly running log distinct from the machine CSV: whole-dollar money, explicit per-leg
columns (up to 3 legs — markets are at most 3-way), and a ``new_or_repeat`` flag computed by
comparing each arb's structural ``arb_id`` against the set seen in the immediately previous scan.

openpyxl is required (see requirements.txt). Failures never crash the run — they are logged.
"""
from __future__ import annotations

import logging
import os
import re
from typing import Any, Optional

MAX_LEGS = 3

# A structural arb_id is a 16-char lowercase hex hash. Used to detect the legacy column shift.
_HEX16 = re.compile(r"^[0-9a-f]{16}$")

# Fixed column order. Per-leg columns repeat for leg 1..MAX_LEGS.
_BASE_HEAD = ["scan_time_local", "scan_time_utc", "game", "kickoff_time", "tournament", "market"]
_LEG_HEAD = ["book", "outcome", "odds", "limit", "stake"]
_TAIL_HEAD = ["S", "ROI_pct", "T_max", "total_investment", "guaranteed_profit", "type",
              "low_confidence", "unverified_limit_books", "capital_lockup_days", "new_or_repeat", "arb_id"]


def columns() -> list[str]:
    head = list(_BASE_HEAD)
    for i in range(1, MAX_LEGS + 1):
        head += [f"leg{i}_{c}" for c in _LEG_HEAD]
    return head + _TAIL_HEAD


def _flatten(row: dict[str, Any]) -> dict[str, Any]:
    """Turn one structured arb row (with a ``legs`` list) into the flat leg1.. column dict."""
    flat = {k: row.get(k) for k in _BASE_HEAD}
    legs = row.get("legs", [])
    for i in range(1, MAX_LEGS + 1):
        leg = legs[i - 1] if i - 1 < len(legs) else {}
        for c in _LEG_HEAD:
            flat[f"leg{i}_{c}"] = leg.get(c) if leg else None
    for k in _TAIL_HEAD:
        flat[k] = row.get(k)
    return flat


def _prev_scan_arb_ids(ws, cols: list[str]) -> set[str]:
    """The set of arb_ids belonging to the most recent scan already in the sheet.

    'Most recent scan' = all rows sharing the maximum scan_time_utc value present. Used so a row
    written this scan is REPEAT iff its arb_id was in the previous scan, else NEW.
    """
    try:
        utc_i = cols.index("scan_time_utc")
        id_i = cols.index("arb_id")
    except ValueError:  # pragma: no cover - header mismatch
        return set()
    latest: Optional[str] = None
    ids_by_scan: dict[str, set[str]] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or len(r) <= max(utc_i, id_i):
            continue
        scan_utc, arb_id = r[utc_i], r[id_i]
        if scan_utc is None or arb_id is None:
            continue
        scan_utc, arb_id = str(scan_utc), str(arb_id)
        ids_by_scan.setdefault(scan_utc, set()).add(arb_id)
        if latest is None or scan_utc > latest:
            latest = scan_utc
    return ids_by_scan.get(latest, set()) if latest is not None else set()


def _read_header(ws) -> list[Any]:
    """The first row's cell values (raw, including any trailing None/empty cells)."""
    if ws.max_row < 1:
        return []
    return [c.value for c in ws[1]]


def _is_hex16(v: Any) -> bool:
    return v is not None and bool(_HEX16.match(str(v)))


def migrate_sheet(ws, log: Optional[logging.Logger] = None) -> int:
    """Rewrite ``ws`` in place so its header and rows match the current ``columns()``.

    Every existing data row is re-mapped to the current schema *by column name*, never by
    position, so adding/removing/reordering columns can never silently shift values again.

    Special-cases the known legacy shift: when ``capital_lockup_days`` was inserted before
    ``new_or_repeat``/``arb_id`` but the workbook header was never migrated, rows written after
    the change land one column to the right — the real arb_id sits in a trailing unnamed column
    while the "arb_id" column holds the NEW/REPEAT flag. Detect this (arb_id cell not a 16-hex
    hash while the trailing cell is) and un-swap them, leaving ``capital_lockup_days`` blank for
    those legacy rows.

    Returns the number of rows whose arb_id/new_or_repeat were un-swapped (repaired).
    """
    cols = columns()
    header = _read_header(ws)

    # Map header name -> first column index that carries it; track trailing unnamed columns.
    name_to_idx: dict[str, int] = {}
    for i, h in enumerate(header):
        if h is not None and h != "":
            name_to_idx.setdefault(str(h), i)
    arb_idx = name_to_idx.get("arb_id")
    unnamed_idxs = [i for i, h in enumerate(header) if h is None or h == ""]
    trailing_idx = unnamed_idxs[-1] if unnamed_idxs else None

    data = list(ws.iter_rows(min_row=2, values_only=True))
    new_rows: list[list[Any]] = []
    repaired = 0
    for r in data:
        # Name -> value using the on-disk header positions.
        rec: dict[str, Any] = {}
        for i, h in enumerate(header):
            if h is not None and h != "" and i < len(r):
                rec[str(h)] = r[i]

        arb_val = r[arb_idx] if (arb_idx is not None and arb_idx < len(r)) else None
        trail_val = r[trailing_idx] if (trailing_idx is not None and trailing_idx < len(r)) else None

        if not _is_hex16(arb_val) and _is_hex16(trail_val):
            # Legacy shift: real arb_id is in the trailing column; the "arb_id" column holds
            # the NEW/REPEAT flag. Un-swap and blank out capital_lockup_days for these rows.
            rec["arb_id"] = trail_val
            rec["new_or_repeat"] = arb_val
            rec["capital_lockup_days"] = None
            repaired += 1

        new_rows.append([rec.get(c) for c in cols])

    # Rewrite the whole sheet (clears any extra/trailing columns too).
    if ws.max_row >= 1:
        ws.delete_rows(1, ws.max_row)
    ws.append(cols)
    for nr in new_rows:
        ws.append(nr)

    if log is not None:
        log.info("excel_log: migrated sheet to current schema (%d data rows, %d legacy row(s) repaired).",
                 len(new_rows), repaired)
    return repaired


_SCOREBOARD_SHEET = "shadow_scoreboard"
_SB_COLS = ["rank", "book", "unlock_count", "total_appearances", "distinct_games",
            "avg_roi_pct", "best_roi_pct", "window_hours", "updated_utc"]


def write_scoreboard(path: str, rows: list[dict[str, Any]], window_hours: float,
                     updated_utc: str, log: logging.Logger) -> int:
    """(Re)write the 'shadow_scoreboard' sheet — fully refreshed each scan, ranked rows already
    sorted by the caller. Creates the workbook (with an empty 'arbs' sheet) if it does not exist yet.
    Returns rows written (0 on error)."""
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:  # pragma: no cover - dependency missing
        log.error("openpyxl not installed — cannot write scoreboard to %s.", path)
        return 0
    try:
        if os.path.exists(path):
            wb = load_workbook(path)
        else:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            wb = Workbook()
            wb.active.title = "arbs"          # keep the arbs sheet present + headed for consistency
            wb.active.append(columns())
        if _SCOREBOARD_SHEET in wb.sheetnames:
            del wb[_SCOREBOARD_SHEET]
        ws = wb.create_sheet(_SCOREBOARD_SHEET)
        ws.append(_SB_COLS)
        for i, r in enumerate(rows, 1):
            ws.append([i, r["book"], r["unlock_count"], r["total_appearances"],
                       r["distinct_games"], r["avg_roi_pct"], r["best_roi_pct"],
                       int(window_hours), updated_utc])
        wb.save(path)
        return len(rows)
    except Exception as exc:  # pragma: no cover - disk / openpyxl error
        log.error("Failed to write scoreboard sheet to %s: %s", path, exc)
        return 0


def append_arbs(path: str, rows: list[dict[str, Any]], log: logging.Logger) -> int:
    """Append ``rows`` (structured arb dicts) to the xlsx, creating it with a header if missing.

    Sets each row's ``new_or_repeat`` by comparing its ``arb_id`` to the previous scan's ids.
    Returns the number of rows written (0 on no rows or on error)."""
    if not rows:
        return 0
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:  # pragma: no cover - dependency missing
        log.error("openpyxl not installed — cannot write %s (pip install openpyxl).", path)
        return 0

    cols = columns()
    try:
        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb.active
            # Startup integrity guard: the on-disk header MUST match the current schema before we
            # append, or rows land in the wrong columns. If it drifted, migrate once, in place.
            header = _read_header(ws)
            if not header:
                ws.append(cols)
            elif header != cols:
                log.info("excel_log: header drift in %s (on-disk header != columns()) — migrating.", path)
                repaired = migrate_sheet(ws, log)
                log.info("excel_log: migration of %s repaired %d misaligned legacy row(s).", path, repaired)
        else:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            wb = Workbook()
            ws = wb.active
            ws.title = "arbs"
            ws.append(cols)

        # Hard guarantee before any append: never write a misaligned row again.
        assert _read_header(ws) == cols, "excel_log: header still mismatched after migration"

        prev_ids = _prev_scan_arb_ids(ws, cols)
        for row in rows:
            row["new_or_repeat"] = "REPEAT" if str(row.get("arb_id")) in prev_ids else "NEW"
            flat = _flatten(row)
            ws.append([flat.get(c) for c in cols])

        wb.save(path)
        return len(rows)
    except Exception as exc:  # pragma: no cover - disk / openpyxl error
        log.error("Failed to write xlsx log %s: %s", path, exc)
        return 0
