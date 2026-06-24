"""Tests for the Excel running log (src/excel_log.py): header, flatten, new_or_repeat."""
from __future__ import annotations

import os

from openpyxl import load_workbook

from src import excel_log


def _row(arb_id, scan_utc, game="A vs B", legs=None):
    return {
        "scan_time_local": "2026-06-17T10:00:00-04:00",
        "scan_time_utc": scan_utc,
        "game": game,
        "kickoff_time": "2026-06-17T15:00:00-04:00",
        "tournament": "World Cup",
        "market": "1x2 Full Time Result",
        "legs": legs or [
            {"book": "pinnacle", "outcome": "A", "odds": 2.10, "limit": 1500, "stake": 700},
            {"book": "1xbet", "outcome": "B", "odds": 2.05, "limit": "UNVERIFIED", "stake": 716},
        ],
        "S": 0.964, "ROI_pct": 3.73, "T_max": 1416, "total_investment": 1416,
        "guaranteed_profit": 53, "type": "REAL", "low_confidence": "N",
        "unverified_limit_books": "1xbet", "arb_id": arb_id,
    }


def test_creates_file_with_header_and_flattens_legs(tmp_path):
    path = str(tmp_path / "arbs_log.xlsx")
    n = excel_log.append_arbs(path, [_row("id1", "2026-06-17T10:00:00Z")], _Log())
    assert n == 1 and os.path.exists(path)

    wb = load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == excel_log.columns()
    assert "leg1_book" in header and "leg2_limit" in header and "arb_id" in header

    row = {h: v for h, v in zip(header, [c.value for c in ws[2]])}
    assert row["game"] == "A vs B"
    assert row["leg1_book"] == "pinnacle" and row["leg1_stake"] == 700
    assert row["leg2_limit"] == "UNVERIFIED"            # string preserved
    assert row["leg3_book"] is None                      # only 2 legs -> leg3 empty
    assert row["new_or_repeat"] == "NEW"                 # nothing before it


def test_new_or_repeat_against_previous_scan(tmp_path):
    path = str(tmp_path / "arbs_log.xlsx")
    log = _Log()
    # Scan 1: two arbs.
    excel_log.append_arbs(path, [_row("keep", "2026-06-17T10:00:00Z"),
                                 _row("gone", "2026-06-17T10:00:00Z")], log)
    # Scan 2: one repeat (keep) + one brand-new (fresh).
    excel_log.append_arbs(path, [_row("keep", "2026-06-17T10:15:00Z"),
                                 _row("fresh", "2026-06-17T10:15:00Z")], log)

    wb = load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    rows = [{h: v for h, v in zip(header, [c.value for c in r])} for r in ws.iter_rows(min_row=2)]
    scan2 = {r["arb_id"]: r["new_or_repeat"] for r in rows if r["scan_time_utc"] == "2026-06-17T10:15:00Z"}
    assert scan2 == {"keep": "REPEAT", "fresh": "NEW"}


# ---------------------------------------------------------------------------
# Schema-drift / migration tests
# ---------------------------------------------------------------------------

# The header the workbook was created with BEFORE capital_lockup_days was added: the old tail
# (no capital_lockup_days) plus a trailing empty column. This is the exact shape that caused the
# column-shift bug.
_OLD_TAIL = ["S", "ROI_pct", "T_max", "total_investment", "guaranteed_profit", "type",
             "low_confidence", "unverified_limit_books", "new_or_repeat", "arb_id"]


def _old_header():
    head = list(excel_log._BASE_HEAD)
    for i in range(1, excel_log.MAX_LEGS + 1):
        head += [f"leg{i}_{c}" for c in excel_log._LEG_HEAD]
    return head + _OLD_TAIL + [None]   # trailing empty column


def _make_legacy_workbook(path):
    """A workbook with the OLD header and one row written the buggy (shifted) way:
    the 'arb_id' column holds 'NEW' and the trailing unnamed column holds the real hash."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "arbs"
    old = _old_header()
    ws.append(old)

    # Build a base record by name, then write it shifted: capital_lockup_days never existed, so
    # new_or_repeat slid into the 'arb_id' column and arb_id into the trailing unnamed column.
    base = {
        "scan_time_local": "2026-06-19T14:02:00-04:00", "scan_time_utc": "2026-06-19T18:02:00Z",
        "game": "X vs Y", "kickoff_time": "2026-06-20T15:00:00-04:00", "tournament": "World Cup",
        "market": "Full Time Result",
        "leg1_book": "pinnacle", "leg1_outcome": "X", "leg1_odds": 2.0, "leg1_limit": 1000, "leg1_stake": 500,
        "leg2_book": "1xbet", "leg2_outcome": "Y", "leg2_odds": 2.1, "leg2_limit": "UNVERIFIED", "leg2_stake": 476,
        "S": 0.97, "ROI_pct": 3.0, "T_max": 976, "total_investment": 976,
        "guaranteed_profit": 24, "type": "SHADOW", "low_confidence": "N", "unverified_limit_books": "1xbet",
    }
    row = [base.get(h) for h in old]            # everything up to unverified_limit_books by name
    nr_idx = old.index("new_or_repeat")
    arb_idx = old.index("arb_id")
    trail_idx = len(old) - 1                     # the trailing None column
    row[nr_idx] = None                           # capital_lockup_days value landed here (blank)
    row[arb_idx] = "REPEAT"                       # the NEW/REPEAT flag landed in 'arb_id'
    row[trail_idx] = "abcdef0123456789"          # the real 16-hex arb_id landed in the trailing col
    ws.append(row)
    wb.save(path)
    return base


def test_append_to_stale_header_triggers_migration(tmp_path):
    path = str(tmp_path / "arbs_log.xlsx")
    _make_legacy_workbook(path)

    # Appending a fresh row must first migrate the sheet, then land everything in the right columns.
    n = excel_log.append_arbs(path, [_row("ffffffffffffffff", "2026-06-20T10:00:00Z", game="New vs Row")], _Log())
    assert n == 1

    wb = load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == excel_log.columns()          # header rewritten to current schema

    rows = [{h: v for h, v in zip(header, [c.value for c in r])} for r in ws.iter_rows(min_row=2)]
    new = next(r for r in rows if r["game"] == "New vs Row")
    assert new["arb_id"] == "ffffffffffffffff"
    assert new["new_or_repeat"] == "NEW"
    assert new["capital_lockup_days"] is None     # not supplied -> blank, not shifted


def test_migration_unswaps_legacy_arb_id_and_new_or_repeat(tmp_path):
    path = str(tmp_path / "arbs_log.xlsx")
    base = _make_legacy_workbook(path)

    wb = load_workbook(path)
    repaired = excel_log.migrate_sheet(wb.active)
    assert repaired == 1

    header = [c.value for c in wb.active[1]]
    assert header == excel_log.columns()
    row = {h: v for h, v in zip(header, [c.value for c in wb.active[2]])}

    # The known legacy row is restored: real hash back in arb_id, flag back in new_or_repeat.
    assert row["arb_id"] == "abcdef0123456789"
    assert row["new_or_repeat"] == "REPEAT"
    assert row["capital_lockup_days"] is None
    # And the rest of the row stayed put (remapped by name).
    assert row["game"] == base["game"]
    assert row["leg1_book"] == "pinnacle"
    assert row["unverified_limit_books"] == "1xbet"
    assert row["leg2_limit"] == "UNVERIFIED"


def test_fresh_workbook_is_unaffected_by_guard(tmp_path):
    path = str(tmp_path / "arbs_log.xlsx")
    # First append creates a current-schema workbook; a second append must not migrate or shift.
    excel_log.append_arbs(path, [_row("aaaaaaaaaaaaaaaa", "2026-06-20T10:00:00Z")], _Log())
    excel_log.append_arbs(path, [_row("bbbbbbbbbbbbbbbb", "2026-06-20T10:15:00Z")], _Log())

    wb = load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == excel_log.columns()
    rows = [{h: v for h, v in zip(header, [c.value for c in r])} for r in ws.iter_rows(min_row=2)]
    assert len(rows) == 2
    for r in rows:
        # arb_id stayed a valid hash in its own column; no row got shifted by a spurious migration.
        assert excel_log._is_hex16(r["arb_id"])
        assert r["new_or_repeat"] in ("NEW", "REPEAT")


class _Log:
    def info(self, *a, **k):
        pass

    def error(self, *a, **k):
        pass
