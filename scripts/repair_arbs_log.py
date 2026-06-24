"""One-shot repair for data/arbs_log.xlsx column-shift bug.

When ``capital_lockup_days`` was added to ``excel_log.columns()`` but the existing workbook kept
its old header (plus a trailing empty column), every row appended afterwards was written one
column to the right: the "arb_id" column holds the NEW/REPEAT flag and the real 16-hex arb_id
sits in the trailing unnamed column.

This script loads the workbook, backs the original up to arbs_log.bak.xlsx, applies the same
name-based remap + arb_id/new_or_repeat un-swap used by ``excel_log.migrate_sheet`` (so the live
appender and this one-shot stay in lockstep), and saves the corrected file. It prints a
before/after count of rows whose arb_id is a valid 16-hex hash.

Usage:  python scripts/repair_arbs_log.py [path/to/arbs_log.xlsx]
"""
from __future__ import annotations

import os
import shutil
import sys

# Allow running as a plain script (python scripts/repair_arbs_log.py).
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook

from src import excel_log

DEFAULT_PATH = os.path.join("data", "arbs_log.xlsx")


def _count_valid_arb_ids(ws) -> int:
    """Count data rows whose 'arb_id' header column holds a valid 16-hex hash."""
    header = excel_log._read_header(ws)
    if "arb_id" not in [str(h) for h in header if h is not None]:
        return 0
    idx = [str(h) if h is not None else None for h in header].index("arb_id")
    n = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        v = r[idx] if idx < len(r) else None
        if excel_log._is_hex16(v):
            n += 1
    return n


def main(path: str = DEFAULT_PATH) -> int:
    if not os.path.exists(path):
        print(f"ERROR: {path} does not exist.")
        return 1

    wb = load_workbook(path)
    ws = wb.active

    before = _count_valid_arb_ids(ws)
    total = max(ws.max_row - 1, 0)
    print(f"Loaded {path}: {total} data rows; valid 16-hex arb_id before: {before}/{total}")

    backup = os.path.join(os.path.dirname(path), "arbs_log.bak.xlsx")
    shutil.copyfile(path, backup)
    print(f"Backed up original to {backup}")

    repaired = excel_log.migrate_sheet(ws)
    after = _count_valid_arb_ids(ws)
    print(f"Migration repaired {repaired} legacy row(s) (arb_id/new_or_repeat un-swapped).")

    wb.save(path)
    print(f"Saved corrected workbook to {path}; valid 16-hex arb_id after: {after}/{total}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PATH))
